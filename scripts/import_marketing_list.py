#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导入《2026年客户营销清单（安徽省分公司）.xlsx》到 visitpro 数据库。
三个 sheet（地方政府/金融机构/产业客户）分别映射为对应 clientType 的客户记录，
类型专属字段 + 协议/项目信息存入 typeProfile JSON 列。

用法: python3 scripts/import_marketing_list.py <xlsx路径> [--replace]
      --replace: 先删除本脚本生成的记录（id 前缀 imp-gov/imp-fin/imp-ent）再导入
"""
import sys
import json
import re
from datetime import datetime, date

import openpyxl

STATUS_LEAD = '潜在客户'


def esc(v):
    """SQL 字符串转义"""
    if v is None:
        return 'NULL'
    s = str(v).replace('\\', '\\\\').replace("'", "\\'")
    return "'" + s + "'"


def clean_text(v):
    if v is None:
        return ''
    # 去掉零宽字符与多余空白
    return re.sub(r'[\u200b\u200c\u200d\ufeff\xa0]', '', str(v)).strip()


def fmt_date(v):
    if v is None or v == '':
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime('%Y-%m-%d')
    s = clean_text(v)
    return s or None


def fmt_ratio(v):
    """持股比例：0.5859 -> 58.59%；文本如 '100%（国有独资）' 原样保留"""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        pct = v * 100
        return ('%g' % round(pct, 2)) + '%'
    return clean_text(v) or None


def yes(v):
    return clean_text(v) == '是'


def agreement(row):
    """row: dict 列字母->值, 按 sheet 传入签署四列"""
    return row.get('agreement') or {}


def insert_sql(cid, name, industry, region, client_type, tp):
    tp_json = json.dumps(tp, ensure_ascii=False)
    return (
        "INSERT INTO clients (id, name, industry, status, clientType, region, "
        "contacts, customFields, typeProfile, tags) VALUES "
        f"({esc(cid)}, {esc(name)}, {esc(industry)}, {esc(STATUS_LEAD)}, "
        f"{esc(client_type)}, {esc(region)}, '[]', '{{}}', {esc(tp_json)}, '[]') "
        f"ON DUPLICATE KEY UPDATE name=VALUES(name), industry=VALUES(industry), "
        f"clientType=VALUES(clientType), region=VALUES(region), typeProfile=VALUES(typeProfile);"
    )


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    path = sys.argv[1]
    replace = '--replace' in sys.argv

    wb = openpyxl.load_workbook(path, data_only=True)
    stmts = []
    if replace:
        stmts.append("DELETE FROM clients WHERE id LIKE 'imp-gov-%' OR id LIKE 'imp-fin-%' OR id LIKE 'imp-ent-%';")

    # ---------- 地方政府 ----------
    ws = wb['地方政府']
    count_gov = 0
    for i in range(3, ws.max_row + 1):
        name = clean_text(ws.cell(row=i, column=2).value)
        if not name:
            continue
        count_gov += 1
        signed = yes(ws.cell(row=i, column=5).value)
        landed = yes(ws.cell(row=i, column=9).value)
        tp = {
            'adminLevel': clean_text(ws.cell(row=i, column=3).value) or None,
            'reportingUnit': clean_text(ws.cell(row=i, column=13).value) or None,
            'agreement': {
                'signed': signed,
                'party': clean_text(ws.cell(row=i, column=6).value) or None,
                'signDate': fmt_date(ws.cell(row=i, column=7).value),
                'expireDate': fmt_date(ws.cell(row=i, column=8).value),
            },
            'project': {
                'landed': landed,
                'projectNo': clean_text(ws.cell(row=i, column=10).value) or None,
                'projectName': clean_text(ws.cell(row=i, column=11).value) or None,
                'scale': ws.cell(row=i, column=12).value if isinstance(ws.cell(row=i, column=12).value, (int, float)) else None,
            },
        }
        stmts.append(insert_sql(f'imp-gov-{count_gov}', name, '', clean_text(ws.cell(row=i, column=4).value), '地方政府', tp))

    # ---------- 金融机构 ----------
    ws = wb['金融机构']
    count_fin = 0
    for i in range(3, ws.max_row + 1):
        name = clean_text(ws.cell(row=i, column=2).value)
        if not name:
            continue
        count_fin += 1
        signed = yes(ws.cell(row=i, column=12).value)
        landed = yes(ws.cell(row=i, column=16).value)
        scale = ws.cell(row=i, column=19).value
        tp = {
            'creditCode': clean_text(ws.cell(row=i, column=3).value) or None,
            'stockCode': clean_text(ws.cell(row=i, column=4).value) if clean_text(ws.cell(row=i, column=4).value) not in ('', '-') else None,
            'foundedDate': fmt_date(ws.cell(row=i, column=5).value),
            'majorShareholder': clean_text(ws.cell(row=i, column=6).value) or None,
            'majorShareholderRatio': fmt_ratio(ws.cell(row=i, column=7).value),
            'finCategory': clean_text(ws.cell(row=i, column=9).value) or None,
            'finSubCategory': clean_text(ws.cell(row=i, column=10).value) or None,
            'finRank': clean_text(ws.cell(row=i, column=11).value) or None,
            'reportingUnit': clean_text(ws.cell(row=i, column=20).value) or None,
            'agreement': {
                'signed': signed,
                'party': clean_text(ws.cell(row=i, column=13).value) or None,
                'signDate': fmt_date(ws.cell(row=i, column=14).value),
                'expireDate': fmt_date(ws.cell(row=i, column=15).value),
            },
            'project': {
                'landed': landed,
                'projectNo': clean_text(ws.cell(row=i, column=17).value) or None,
                'projectName': clean_text(ws.cell(row=i, column=18).value) or None,
                'scale': scale if isinstance(scale, (int, float)) else None,
            },
        }
        stmts.append(insert_sql(f'imp-fin-{count_fin}', name, '金融业', clean_text(ws.cell(row=i, column=8).value), '金融机构', tp))

    # ---------- 产业客户 ----------
    ws = wb['产业客户']
    count_ent = 0
    for i in range(3, ws.max_row + 1):
        name = clean_text(ws.cell(row=i, column=2).value)
        if not name:
            continue
        count_ent += 1
        signed = yes(ws.cell(row=i, column=16).value)
        landed = yes(ws.cell(row=i, column=20).value)
        scale = ws.cell(row=i, column=23).value
        industry = clean_text(ws.cell(row=i, column=11).value)
        tp = {
            'creditCode': clean_text(ws.cell(row=i, column=3).value) or None,
            'stockCode': clean_text(ws.cell(row=i, column=4).value) if clean_text(ws.cell(row=i, column=4).value) not in ('', '-') else None,
            'foundedDate': fmt_date(ws.cell(row=i, column=5).value),
            'majorShareholder': clean_text(ws.cell(row=i, column=6).value) or None,
            'majorShareholderRatio': fmt_ratio(ws.cell(row=i, column=7).value),
            'groupOwner': clean_text(ws.cell(row=i, column=9).value) or None,
            'entCategory': clean_text(ws.cell(row=i, column=10).value) or None,
            'industryCategory': industry or None,
            'industrySub': clean_text(ws.cell(row=i, column=12).value) or None,
            'industryCode': clean_text(ws.cell(row=i, column=13).value) or None,
            'creditRating': clean_text(ws.cell(row=i, column=14).value) if clean_text(ws.cell(row=i, column=14).value) not in ('', '-') else None,
            'top500Rank': clean_text(ws.cell(row=i, column=15).value) if clean_text(ws.cell(row=i, column=15).value) not in ('', '-') else None,
            'reportingUnit': clean_text(ws.cell(row=i, column=24).value) or None,
            'agreement': {
                'signed': signed,
                'party': clean_text(ws.cell(row=i, column=17).value) or None,
                'signDate': fmt_date(ws.cell(row=i, column=18).value),
                'expireDate': fmt_date(ws.cell(row=i, column=19).value),
            },
            'project': {
                'landed': landed,
                'projectNo': clean_text(ws.cell(row=i, column=21).value) or None,
                'projectName': clean_text(ws.cell(row=i, column=22).value) or None,
                'scale': scale if isinstance(scale, (int, float)) else None,
            },
        }
        stmts.append(insert_sql(f'imp-ent-{count_ent}', name, industry, clean_text(ws.cell(row=i, column=8).value), '产业客户', tp))

    out = '\n'.join(stmts)
    with open('/tmp/import_marketing_list.sql', 'w', encoding='utf-8') as f:
        f.write("SET NAMES utf8mb4;\nUSE visitpro;\n" + out + '\n')

    print(f'地方政府 {count_gov} 条, 金融机构 {count_fin} 条, 产业客户 {count_ent} 条')
    print('SQL 已写入 /tmp/import_marketing_list.sql')


if __name__ == '__main__':
    main()
