#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
导入《安徽省分公司通讯录（2026年7月）.xlsx》到 visitpro.users 表。
- Sheet1 为通讯录（B/C 列部门合并单元格），按部门段归属人员
- 领导班子（无部门段）归入"总经理室"
- 物业人员段按需求跳过
- 邮箱用 {手机号}@visitpro.com 生成，默认密码 Visit@2026（SHA-256）

用法: python3 scripts/import_address_book.py <xlsx路径> [--replace]
      --replace: 先删除本脚本生成的用户（id 前缀 user_imp_）再导入
"""
import sys
import json
import re
import hashlib

import openpyxl

DEFAULT_PASSWORD = 'Visit@2026'
PASSWORD_HASH = hashlib.sha256(DEFAULT_PASSWORD.encode()).hexdigest()
SKIP_DEPTS = {'物业人员'}
LEADER_DEPT = '总经理室'


def esc(v):
    if v is None:
        return 'NULL'
    s = str(v).replace('\\', '\\\\').replace("'", "\\'")
    return "'" + s + "'"


def clean_name(v):
    """去除姓名中的空格/全角空格"""
    return re.sub(r'[\s\u3000]+', '', str(v)).strip()


def clean_text(v):
    if v is None:
        return ''
    return re.sub(r'[\u200b\u200c\u200d\ufeff\xa0]', '', str(v)).replace('\n', ' ').strip()


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    path = sys.argv[1]
    replace = '--replace' in sys.argv

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Sheet1']

    # 合并单元格 -> 左上角值
    merge_map = {}
    for rng in ws.merged_cells.ranges:
        top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merge_map[(r, c)] = top_left

    def val(r, c):
        v = ws.cell(row=r, column=c).value
        if v is None or str(v).strip() == '':
            v = merge_map.get((r, c))
        return str(v).strip() if v is not None else ''

    users = []       # (序号, 部门名, 姓名, 职务, 办公电话, 手机)
    skipped = 0
    for i in range(4, ws.max_row + 1):
        dept = val(i, 2) or val(i, 3)
        name = clean_name(val(i, 4)) if val(i, 4) else ''
        if not name:
            continue
        if dept in SKIP_DEPTS:
            skipped += 1
            continue
        position = clean_text(val(i, 5))
        office_phone = clean_text(val(i, 6))
        mobile = re.sub(r'[\s-]', '', val(i, 7))
        users.append((dept or LEADER_DEPT, name, position, office_phone, mobile))

    stmts = []
    if replace:
        stmts.append("DELETE FROM users WHERE id LIKE 'user_imp_%';")

    dept_count = {}
    emails = set()
    for idx, (dept, name, position, office_phone, mobile) in enumerate(users, start=1):
        uid = f'user_imp_{idx:03d}'
        email = f'{mobile}@visitpro.com' if mobile else f'staff{idx:03d}@visitpro.com'
        if email in emails:
            print(f'警告: 邮箱重复 {email}（{name}），请检查')
        emails.add(email)
        dept_count[dept] = dept_count.get(dept, 0) + 1
        cf = json.dumps(
            {k: v for k, v in [('position', position), ('officePhone', office_phone)] if v},
            ensure_ascii=False
        )
        dept_sql = f"(SELECT id FROM departments WHERE name = {esc(dept)} LIMIT 1)"
        stmts.append(
            "INSERT INTO users (id, name, email, phone, password, roleId, departmentId, status, customFields) VALUES "
            f"({esc(uid)}, {esc(name)}, {esc(email)}, {esc(mobile or None)}, {esc(PASSWORD_HASH)}, "
            f"'role_staff', {dept_sql}, 'active', {esc(cf)}) "
            f"ON DUPLICATE KEY UPDATE name=VALUES(name), email=VALUES(email), phone=VALUES(phone), "
            f"departmentId=VALUES(departmentId), customFields=VALUES(customFields);"
        )

    with open('/tmp/import_address_book.sql', 'w', encoding='utf-8') as f:
        f.write("SET NAMES utf8mb4;\nUSE visitpro;\n" + '\n'.join(stmts) + '\n')

    print(f'待导入 {len(users)} 人（跳过物业人员 {skipped} 人）')
    for d, n in dept_count.items():
        print(f'  {d}: {n} 人')
    print('SQL 已写入 /tmp/import_address_book.sql')


if __name__ == '__main__':
    main()
